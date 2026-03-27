import os, io, base64, traceback, json, re
from flask import Flask, request, jsonify
from openpyxl import load_workbook

app = Flask(__name__)
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "bov_template.xlsx")

# ── JLT DATABASE ─────────────────────────────────────────────────────────────
JLT_DB_PATH = os.path.join(os.path.dirname(__file__), "jlt_data.json")

# Embedded baseline data (Denver/Aurora/Boulder, CO - JLT July 2025)
EMBEDDED_JLT = [{"name":"Cedar Village Mobile Home Park","address":"15814 East Colfax Avenue","city":"Aurora","state":"CO","zip":"80011","phone":"303-344-3831","spaces":"44","occupancy_pct":"100%","avg_rent":"960","adj_avg_rent":"960","low_rent":"960","high_rent":"960","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Foxridge Farms","address":"26900 East Colfax Avenue","city":"Aurora","state":"CO","zip":"80018","phone":"303-344-0655","spaces":"481","occupancy_pct":"98%","avg_rent":"1085","adj_avg_rent":"","low_rent":"1085","high_rent":"1160","utility":"W/S","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Water, Sewer incl."},{"name":"Friendly Village of Aurora","address":"1711 Roosevelt Way","city":"Aurora","state":"CO","zip":"80011","phone":"303-360-9530","spaces":"439","occupancy_pct":"100%","avg_rent":"1187","adj_avg_rent":"1178","low_rent":"1187","high_rent":"1187","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Green Acres Mobile Home Park","address":"1540 Billings Street","city":"Aurora","state":"CO","zip":"80011","phone":"303-364-3744","spaces":"209","occupancy_pct":"96%","avg_rent":"1295","adj_avg_rent":"1295","low_rent":"1295","high_rent":"1295","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Hillcrest Village","address":"1600 Sable Boulevard","city":"Aurora","state":"CO","zip":"80011","phone":"855-514-1062","spaces":"602","occupancy_pct":"100%","avg_rent":"1227","adj_avg_rent":"","low_rent":"1215","high_rent":"1240","utility":"W/S","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Water, Sewer incl."},{"name":"The Meadows","address":"14470 East 13th Avenue","city":"Aurora","state":"CO","zip":"80011","phone":"303-343-1414","spaces":"303","occupancy_pct":"100%","avg_rent":"1189","adj_avg_rent":"1189","low_rent":"1189","high_rent":"1189","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Woodshire East Mobile Home Community","address":"17501 East Colfax Avenue","city":"Aurora","state":"CO","zip":"80011","phone":"303-341-2354","spaces":"239","occupancy_pct":"100%","avg_rent":"980","adj_avg_rent":"953","low_rent":"980","high_rent":"980","utility":"T","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Trash incl."},{"name":"Arbordale Acres","address":"507 East Spaulding Street","city":"Lafayette","state":"CO","zip":"80026","phone":"303-665-6457","spaces":"286","occupancy_pct":"100%","avg_rent":"1179","adj_avg_rent":"1179","low_rent":"1179","high_rent":"1179","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Boulder Meadows","address":"4500 19Th Street","city":"Boulder","state":"CO","zip":"80304","phone":"303-442-6337","spaces":"638","occupancy_pct":"100%","avg_rent":"971","adj_avg_rent":"950","low_rent":"971","high_rent":"971","utility":"T","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Trash incl."},{"name":"Boulder Ridge Mobile Home Park","address":"11990 E South Boulder Road","city":"Lafayette","state":"CO","zip":"80026","phone":"303-665-6996","spaces":"241","occupancy_pct":"100%","avg_rent":"1060","adj_avg_rent":"","low_rent":"1060","high_rent":"1360","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Eagle Crest","address":"4700 Eagle Crest Boulevard","city":"Firestone","state":"CO","zip":"80504","phone":"888-267-4785","spaces":"441","occupancy_pct":"100%","avg_rent":"1003","adj_avg_rent":"","low_rent":"1003","high_rent":"1034","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Grand Meadow","address":"821 - 17th Avenue","city":"Longmont","state":"CO","zip":"80501","phone":"720-494-0751","spaces":"102","occupancy_pct":"99%","avg_rent":"1047","adj_avg_rent":"","low_rent":"1047","high_rent":"1047","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Lafayette Gardens","address":"11700 South Boulder Road","city":"Lafayette","state":"CO","zip":"80026","phone":"303-665-5822","spaces":"135","occupancy_pct":"100%","avg_rent":"1020","adj_avg_rent":"1020","low_rent":"1020","high_rent":"1020","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"River Valley Village Mobile Home Community","address":"10910 Turner Boulevard","city":"Longmont","state":"CO","zip":"80504","phone":"303-772-3240","spaces":"210","occupancy_pct":"100%","avg_rent":"1030","adj_avg_rent":"","low_rent":"1030","high_rent":"1075","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Vista Village Mobile Home Community","address":"5000 Butte Street","city":"Boulder","state":"CO","zip":"80301","phone":"303-443-3002","spaces":"305","occupancy_pct":"100%","avg_rent":"995","adj_avg_rent":"995","low_rent":"995","high_rent":"995","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Berkeley Village Mobile Home Park","address":"5400 Sheridan Blvd","city":"Arvada","state":"CO","zip":"80002","phone":"303-455-0659","spaces":"395","occupancy_pct":"100%","avg_rent":"1094","adj_avg_rent":"","low_rent":"995","high_rent":"1200","utility":"T","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Trash incl."},{"name":"Casa Estates","address":"860 West 132nd Avenue","city":"Westminster","state":"CO","zip":"80234","phone":"303-452-6792","spaces":"364","occupancy_pct":"100%","avg_rent":"1194","adj_avg_rent":"1194","low_rent":"1194","high_rent":"1194","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Castle Park","address":"2 Darren Street","city":"Castle Rock","state":"CO","zip":"80109","phone":"720-510-3108","spaces":"74","occupancy_pct":"100%","avg_rent":"1060","adj_avg_rent":"1060","low_rent":"1060","high_rent":"1060","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Cimarron Village","address":"12205 North Perry Street","city":"Broomfield","state":"CO","zip":"80020","phone":"855-514-1065","spaces":"327","occupancy_pct":"100%","avg_rent":"1178","adj_avg_rent":"","low_rent":"1178","high_rent":"1200","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Commerce Heights","address":"7701 Brighton Boulevard","city":"Commerce City","state":"CO","zip":"80022","phone":"303-288-2038","spaces":"52","occupancy_pct":"100%","avg_rent":"1035","adj_avg_rent":"","low_rent":"1035","high_rent":"1035","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Countryside Village Denver","address":"9850 Federal Blvd","city":"Federal Heights","state":"CO","zip":"80260","phone":"303-469-1991","spaces":"335","occupancy_pct":"100%","avg_rent":"1143","adj_avg_rent":"","low_rent":"1143","high_rent":"1143","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Denver Cascade","address":"9850 N Federal Boulevard","city":"Federal Heights","state":"CO","zip":"80260","phone":"303-469-1919","spaces":"382","occupancy_pct":"100%","avg_rent":"1295","adj_avg_rent":"1295","low_rent":"1295","high_rent":"1295","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Eastgate Village","address":"3060 East Bridge Street","city":"Brighton","state":"CO","zip":"80601","phone":"303-659-5225","spaces":"452","occupancy_pct":"100%","avg_rent":"1042","adj_avg_rent":"","low_rent":"1019","high_rent":"1079","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Franklin Mobile Home Park","address":"1616 East 78th Avenue","city":"Thornton","state":"CO","zip":"80229","phone":"303-287-4036","spaces":"170","occupancy_pct":"100%","avg_rent":"975","adj_avg_rent":"940","low_rent":"975","high_rent":"975","utility":"T","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Trash incl."},{"name":"Friendly Village Of The Rockies","address":"2100 West 100th Avenue","city":"Thornton","state":"CO","zip":"80260","phone":"303-466-4328","spaces":"524","occupancy_pct":"100%","avg_rent":"1187","adj_avg_rent":"1187","low_rent":"1187","high_rent":"1187","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Front Range","address":"2885 East Midway Boulevard","city":"Denver","state":"CO","zip":"80234","phone":"303-469-1773","spaces":"587","occupancy_pct":"99%","avg_rent":"1193","adj_avg_rent":"1193","low_rent":"1193","high_rent":"1193","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Garden Meadows","address":"6250 North Federal Blvd","city":"Denver","state":"CO","zip":"80221","phone":"303-422-4748","spaces":"100","occupancy_pct":"98%","avg_rent":"1238","adj_avg_rent":"","low_rent":"1238","high_rent":"1238","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Highview Mobile Home Community","address":"8601 Zuni Street","city":"Westminster","state":"CO","zip":"80260","phone":"303-427-3690","spaces":"274","occupancy_pct":"100%","avg_rent":"800","adj_avg_rent":"693","low_rent":"","high_rent":"","utility":"W/S/T","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Water, Sewer, Trash incl."},{"name":"Inspiration Valley","address":"5250 West 53rd Avenue","city":"Arvada","state":"CO","zip":"80002","phone":"303-422-4748","spaces":"140","occupancy_pct":"100%","avg_rent":"1191","adj_avg_rent":"1146","low_rent":"1191","high_rent":"1191","utility":"W","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Water incl."},{"name":"Kimberly Hills","address":"2305 West 92nd Avenue","city":"Federal Heights","state":"CO","zip":"80260","phone":"303-427-6612","spaces":"687","occupancy_pct":"100%","avg_rent":"1183","adj_avg_rent":"1183","low_rent":"1183","high_rent":"1183","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Lamplighter Village","address":"9100 Tejon Street","city":"Federal Heights","state":"CO","zip":"80260","phone":"303-428-3393","spaces":"250","occupancy_pct":"100%","avg_rent":"1183","adj_avg_rent":"1183","low_rent":"1183","high_rent":"1183","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"North County Village","address":"9700 Riverdale Road","city":"Denver","state":"CO","zip":"80229","phone":"303-252-4551","spaces":"425","occupancy_pct":"100%","avg_rent":"1100","adj_avg_rent":"1100","low_rent":"1100","high_rent":"1100","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Shady Lane","address":"6791 Highway 2","city":"Commerce City","state":"CO","zip":"80022","phone":"303-288-2038","spaces":"65","occupancy_pct":"100%","avg_rent":"1012","adj_avg_rent":"1012","low_rent":"1012","high_rent":"1012","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"The Grove at Alta Ridge","address":"1201 W. Thornton Parkway","city":"Thornton","state":"CO","zip":"80260","phone":"877-377-3406","spaces":"409","occupancy_pct":"100%","avg_rent":"1019","adj_avg_rent":"1019","low_rent":"1019","high_rent":"1019","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Thornton Estates","address":"3600 East 88th Avenue","city":"Thornton","state":"CO","zip":"80229","phone":"303-288-0886","spaces":"207","occupancy_pct":"98%","avg_rent":"1124","adj_avg_rent":"1124","low_rent":"1124","high_rent":"1124","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Wikiup","address":"6500 East 88th Avenue","city":"Henderson","state":"CO","zip":"80640","phone":"303-288-2038","spaces":"339","occupancy_pct":"99%","avg_rent":"1193","adj_avg_rent":"1150","low_rent":"1193","high_rent":"1193","utility":"W","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Water incl."},{"name":"Golden Terrace","address":"17801 West Colfax Avenue","city":"Golden","state":"CO","zip":"80401","phone":"303-279-6279","spaces":"311","occupancy_pct":"100%","avg_rent":"1240","adj_avg_rent":"1240","low_rent":"1240","high_rent":"1240","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Mountainside Estates","address":"17190 Mount Vernon Road","city":"Golden","state":"CO","zip":"80401","phone":"303-279-5098","spaces":"226","occupancy_pct":"100%","avg_rent":"1252","adj_avg_rent":"1252","low_rent":"1252","high_rent":"1252","utility":"None","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Tenant-paid"},{"name":"Ridgeline at River Run","address":"3650 S Federal Blvd","city":"Englewood","state":"CO","zip":"80110","phone":"303-761-0121","spaces":"226","occupancy_pct":"96%","avg_rent":"1150","adj_avg_rent":"1055","low_rent":"1150","high_rent":"1150","utility":"W/S/T","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Water, Sewer, Trash incl."},{"name":"Wolhurst Lake","address":"8201 South Sante Fe Drive","city":"Littleton","state":"CO","zip":"80120","phone":"303-795-0777","spaces":"301","occupancy_pct":"100%","avg_rent":"1495","adj_avg_rent":"1382","low_rent":"","high_rent":"","utility":"W/S/T/L","report":"JLT July 2025","market":"Denver/Aurora/Boulder, CO","utility_display":"Water, Sewer, Trash, Lawn incl."}]

def load_jlt_db():
    # Start with embedded data
    db = list(EMBEDDED_JLT)
    # Merge any additional parks from file (uploaded via /parse-jlt)
    if os.path.exists(JLT_DB_PATH):
        try:
            with open(JLT_DB_PATH) as f:
                extra = json.load(f)
            # Add parks not already in embedded set
            embedded_names = {p['name'].lower() for p in db}
            for p in extra:
                if p['name'].lower() not in embedded_names:
                    db.append(p)
        except:
            pass
    return db

def save_jlt_db(parks):
    # Only save non-embedded parks to file
    embedded_names = {p['name'].lower() for p in EMBEDDED_JLT}
    extra = [p for p in parks if p['name'].lower() not in embedded_names]
    try:
        with open(JLT_DB_PATH, 'w') as f:
            json.dump(extra, f, indent=2)
    except:
        pass  # Render filesystem may be read-only

def normalize(s):
    """Normalize string for fuzzy matching"""
    s = s.lower()
    for w in ['mobile home park','mobile home community','manufactured home park',
              'manufactured housing community','mhp','mhc','village','estates',
              'community','park','manor','court','acres','heights','ridge','valley',
              'meadows','pines','oaks','hills','lakes','terrace','gardens']:
        s = s.replace(w, '')
    return re.sub(r'[^a-z0-9\s]', '', s).strip()

def match_score(name1, name2, city1='', city2=''):
    """Return 0-100 match score between two park names"""
    n1, n2 = normalize(name1), normalize(name2)
    # Exact match
    if n1 == n2: return 100
    # One contains the other (only if normalized name has 2+ words to avoid single-word false matches)
    shorter = n1 if len(n1) <= len(n2) else n2
    if (n1 in n2 or n2 in n1) and len(shorter.split()) >= 2:
        return 85
    # Word overlap
    words1 = set(n1.split())
    words2 = set(n2.split())
    if not words1 or not words2: return 0
    overlap = len(words1 & words2)
    score = int(overlap / max(len(words1), len(words2)) * 70)
    # City match bonus
    if city1 and city2 and city1.lower().strip() == city2.lower().strip():
        score += 15
    return min(score, 99)

def find_jlt_match(park_name, city='', state=''):
    """Find best JLT match for a given park name/location"""
    db = load_jlt_db()
    best_score = 0
    best_match = None
    for p in db:
        # Hard-exclude cross-state matches when state is known
        if state and p.get('state','') and state.upper() != p['state'].upper():
            continue
        # Hard-exclude cross-city matches when city is known
        if city and p.get('city','') and city.lower().strip() != p['city'].lower().strip():
            continue
        score = match_score(park_name, p['name'], city, p.get('city',''))
        if score > best_score:
            best_score = score
            best_match = p
    if best_score >= 75:
        return best_match, best_score
    return None, 0

# ── CORS ──────────────────────────────────────────────────────────────────────
@app.after_request
def add_cors(r):
    r.headers["Access-Control-Allow-Origin"] = "*"
    r.headers["Access-Control-Allow-Headers"] = "Content-Type"
    r.headers["Access-Control-Allow-Methods"] = "POST, OPTIONS, GET"
    return r

# ── TEST ──────────────────────────────────────────────────────────────────────
@app.route("/test", methods=["GET"])
def test():
    try:
        wb = load_workbook(TEMPLATE_PATH)
        db = load_jlt_db()
        return jsonify({"status": "ok", "sheets": wb.sheetnames, "jlt_parks": len(db)})
    except Exception as e:
        return jsonify({"status": "error", "error": str(e)})

# ── JLT STATUS ────────────────────────────────────────────────────────────────
@app.route("/jlt-status", methods=["GET"])
def jlt_status():
    db = load_jlt_db()
    markets = {}
    for p in db:
        m = p.get('market', 'Unknown')
        markets[m] = markets.get(m, 0) + 1
    return jsonify({"total_parks": len(db), "markets": markets})

# ── UPLOAD JLT PDF ────────────────────────────────────────────────────────────
@app.route("/upload-jlt", methods=["POST", "OPTIONS"])
def upload_jlt():
    if request.method == "OPTIONS": return "", 200
    try:
        data = request.get_json(force=True)
        new_parks = data.get("parks", [])
        if not new_parks:
            return jsonify({"success": False, "error": "No parks provided"}), 400

        db = load_jlt_db()
        added, updated = 0, 0
        for np in new_parks:
            # Check if already exists
            found = False
            for i, ep in enumerate(db):
                score = match_score(np['name'], ep['name'], np.get('city',''), ep.get('city',''))
                if score >= 85:
                    db[i] = np  # update
                    updated += 1
                    found = True
                    break
            if not found:
                db.append(np)
                added += 1

        save_jlt_db(db)
        return jsonify({"success": True, "added": added, "updated": updated, "total": len(db)})
    except Exception as e:
        return jsonify({"success": False, "error": str(e), "trace": traceback.format_exc()}), 500

# ── RESEARCH RENTS ────────────────────────────────────────────────────────────
@app.route("/research-rents", methods=["POST", "OPTIONS"])
def research_rents():
    if request.method == "OPTIONS": return "", 200
    try:
        d = request.get_json(force=True)
        parks = d.get("parks", [])
        if not parks:
            return jsonify({"success": False, "error": "No parks provided"}), 400

        results = []
        needs_ai = []

        # Step 1: Check JLT database first
        for park in parks:
            name = park.get('name', '')
            city = park.get('city', '')
            state = park.get('state', '')
            # Try to extract city from address
            if not city and park.get('address'):
                addr = park['address']
                parts = addr.split(',')
                if len(parts) >= 2:
                    city = parts[-2].strip()

            jlt_match, score = find_jlt_match(name, city, state)

            if jlt_match:
                results.append({
                    "index": park.get('index', parks.index(park) + 1),
                    "avg_rent": int(jlt_match.get('avg_rent', 0)) if jlt_match.get('avg_rent') else None,
                    "adj_avg_rent": int(jlt_match.get('adj_avg_rent', 0)) if jlt_match.get('adj_avg_rent') else None,
                    "min_rent": int(jlt_match.get('low_rent', 0)) if jlt_match.get('low_rent') else None,
                    "max_rent": int(jlt_match.get('high_rent', 0)) if jlt_match.get('high_rent') else None,
                    "spaces": jlt_match.get('spaces', ''),
                    "occupancy": jlt_match.get('occupancy_pct', ''),
                    "utility": jlt_match.get('utility_display', ''),
                    "source": "JLT " + jlt_match.get('report', 'Report'),
                    "confidence": "high",
                    "match_score": score,
                    "jlt_name": jlt_match['name'],
                    "jlt_city": jlt_match.get('city', '')
                })
            else:
                needs_ai.append(park)

        # Step 2: AI estimates for unmatched parks
        if needs_ai:
            api_key = os.environ.get("ANTHROPIC_API_KEY", "")
            if api_key:
                try:
                    park_list = "\n".join([
                        f"{p.get('index', i+1)}. {p.get('name','')} — {p.get('address','')}"
                        for i, p in enumerate(needs_ai)
                    ])
                    import anthropic
                    client = anthropic.Anthropic(api_key=api_key)
                    message = client.messages.create(
                        model="claude-sonnet-4-20250514",
                        max_tokens=2000,
                        system="""You are a manufactured housing market analyst. Estimate current lot rents for each park based on name and location. Respond ONLY with a JSON array, no markdown:
[{"index":1,"avg_rent":850,"min_rent":800,"max_rent":900,"spaces":120,"utility":"Tenant-paid","source":"AI Estimate","confidence":"medium"}]
confidence: high=specific knowledge, medium=regional estimate, low=rough guess.""",
                        messages=[{"role": "user", "content": f"Estimate lot rents:\n\n{park_list}"}]
                    )
                    text = message.content[0].text
                    s, e = text.find("["), text.rfind("]")
                    if s >= 0 and e > s:
                        ai_results = json.loads(text[s:e+1])
                        # Remap indexes to original park indexes
                        for ar in ai_results:
                            orig_idx = needs_ai[ar['index']-1].get('index') if ar['index']-1 < len(needs_ai) else ar['index']
                            ar['index'] = orig_idx
                            ar['occupancy'] = ar.get('occupancy', '')
                        results.extend(ai_results)
                except Exception as ai_err:
                    print("AI error:", ai_err)
                    # Return basic placeholders for unmatched
                    for p in needs_ai:
                        results.append({
                            "index": p.get('index', 1),
                            "avg_rent": None, "min_rent": None, "max_rent": None,
                            "spaces": "", "utility": "", "source": "Not found",
                            "confidence": "low"
                        })
            else:
                for p in needs_ai:
                    results.append({
                        "index": p.get('index', 1),
                        "avg_rent": None, "source": "No API key", "confidence": "low"
                    })

        # Sort by index
        results.sort(key=lambda x: x.get('index', 999))
        jlt_count = sum(1 for r in results if 'JLT' in r.get('source', ''))
        return jsonify({"success": True, "results": results, "jlt_matches": jlt_count, "ai_estimates": len(results) - jlt_count})

    except Exception as e:
        print("research_rents ERROR:", traceback.format_exc())
        return jsonify({"success": False, "error": str(e), "trace": traceback.format_exc()}), 500

# ── GENERATE BOV ──────────────────────────────────────────────────────────────
@app.route("/generate-bov", methods=["POST", "OPTIONS"])
def generate_bov():
    if request.method == "OPTIONS": return "", 200
    try:
        d = request.get_json(force=True)
        if not d: return jsonify({"success": False, "error": "No JSON received"}), 400
        wb = load_workbook(TEMPLATE_PATH)
        wb.calculation.iterate = True
        wb.calculation.iterateCount = 100
        wb.calculation.iterateDelta = 0.001
        wb["Sales Comps"].sheet_state = "visible"
        wb["Rent Comps"].sheet_state = "visible"
        prop = str(d.get("propName", "Property"))
        today = str(d.get("today", ""))
        year = int(d.get("year", 2026))
        units = int(d.get("units", 0))
        occupied = int(d.get("occupied", units))
        cap_rate = float(d.get("capRate", 0.05))
        def sv(ws, addr, val):
            if val is not None and val != "":
                try: ws[addr].value = val
                except: pass
        ws = wb["BOV Summary"]
        sv(ws,"A3", prop + "  |  " + str(d.get("address","")) + "  |  " + str(units) + " Units  |  Confidential")
        sv(ws,"A4", "Prepared by: Northmarq  |  " + today)
        sv(ws,"C11", prop); sv(ws,"F11", units)
        sv(ws,"C12", "Manufactured Housing Community"); sv(ws,"F12", occupied)
        sv(ws,"C13", str(d.get("address",""))); sv(ws,"C14", str(d.get("rentRange","")))
        sv(ws,"C15", str(d.get("mgmt",""))); sv(ws,"F15", year); sv(ws,"F24", cap_rate)
        ws = wb["Income Statement"]
        sv(ws,"A3", prop + "  |  January - December " + str(year-1) + "  |  Accrual Basis")
        sv(ws,"A4", "Northmarq  |  " + today)
        for addr, key in [
            ("C7","lotRent"),("C8","storageFees"),("C9","appFees"),("C10","lateFees"),
            ("C11","concessions"),("C12","cableIncome"),("C13","miscIncome"),
            ("C15","gasBilled"),("C16","waterBilled"),("C17","sewerBilled"),
            ("C18","garbageBilled"),("C19","electricBilled"),("C27","gasCost"),
            ("C28","waterCost"),("C29","sewerCost"),("C30","electricCost"),("C31","garbageCost"),
            ("C34","advertising"),("C35","travelAuto"),("C36","pestControl"),("C37","landscaping"),
            ("C38","insurance"),("C39","mgrInsurance"),("C40","legalFees"),("C42","poolExpense"),
            ("C43","maintenance"),("C44","cleaning"),("C45","streetRepairs"),("C48","propertyTax"),
            ("C50","officeSupplies"),("C51","internet"),("C53","licensesDues"),
            ("C55","residentMgrSalary"),("C56","rmLabor"),("C57","management"),
            ("C58","payrollTax"),("C59","payrollProcessing"),
        ]:
            val = d.get(key)
            if val:
                try: sv(ws, addr, float(val))
                except: pass
        for row in [7,8,9,10,11,12,13,15,16,17,18,19,27,28,29,30,31,34,35,36,37,38,
                    39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59]:
            for col in ["D","E"]:
                try: ws[col + str(row)].value = None
                except: pass
        ws = wb["Sales Comps"]
        sv(ws,"B3", prop + "  |  Enter comps manually  |  " + today)
        for i, row_data in enumerate(d.get("salesComps",[])[:10]):
            for j, col in enumerate(["B","C","D","E","F","G","H","I","J","K"]):
                if j < len(row_data) and row_data[j] not in [None,"","nan"]:
                    val = row_data[j]
                    if col in ["D","F","G","I"]:
                        try: val = float(str(val).replace(",","").replace("$","").replace("%",""))
                        except: pass
                    try: ws[col + str(7+i)].value = val
                    except: pass
        ws = wb["Rent Comps"]
        sv(ws,"B3", prop + "  |  Enter rent comps manually  |  " + today)
        for i, row_data in enumerate(d.get("rentComps",[])[:10]):
            for j, col in enumerate(["B","C","D","E","F","G","H","I","J"]):
                if j < len(row_data) and row_data[j] not in [None,"","nan"]:
                    val = row_data[j]
                    if col in ["D","E","F","G","I"]:
                        try: val = float(str(val).replace(",","").replace("$","").replace("%",""))
                        except: pass
                    try: ws[col + str(7+i)].value = val
                    except: pass
        ws = wb["5-Year Cash Flow"]
        sv(ws,"A2", prop + "  |  Projected " + str(year) + " - " + str(year+4) + "  |  Capital Markets as of " + today)
        buf = io.BytesIO()
        wb.save(buf)
        return jsonify({"success": True, "b64": base64.b64encode(buf.getvalue()).decode()})
    except Exception as e:
        return jsonify({"success": False, "error": str(e), "trace": traceback.format_exc()}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))

# ── PARSE JLT PDF ─────────────────────────────────────────────────────────────
@app.route("/parse-jlt", methods=["POST", "OPTIONS"])
def parse_jlt():
    if request.method == "OPTIONS": return "", 200
    try:
        d = request.get_json(force=True)
        pdf_b64 = d.get("pdf_b64", "")
        filename = d.get("filename", "report.pdf")
        if not pdf_b64:
            return jsonify({"success": False, "error": "No PDF provided"}), 400

        try:
            import pdfplumber
        except ImportError:
            return jsonify({"success": False, "error": "pdfplumber not installed"}), 500

        pdf_bytes = base64.b64decode(pdf_b64)
        parks = []

        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            # Detect market from first pages
            market = "Unknown Market"
            report_date = "JLT Report"
            for i in range(min(3, len(pdf.pages))):
                t = pdf.pages[i].extract_text() or ''
                if 'JLT Market Report' in t:
                    lines = t.split('\n')
                    for line in lines:
                        if 'CSA' in line or 'Metro' in line or 'Area' in line:
                            if 'JLT' not in line and len(line) < 80:
                                market = line.strip()
                                break
                    # Get date
                    import re as re2
                    date_m = re2.search(r'(January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}', t)
                    if date_m:
                        report_date = 'JLT ' + date_m.group()
                    break

            # Extract individual park pages
            for i in range(len(pdf.pages)):
                page = pdf.pages[i]
                tables = page.extract_tables()
                text = page.extract_text() or ''

                if 'Community Information' not in text or 'Site Info' not in text:
                    continue

                park = {
                    'name': '', 'address': '', 'city': '', 'state': '', 'zip': '',
                    'phone': '', 'spaces': '', 'occupancy_pct': '',
                    'avg_rent': '', 'adj_avg_rent': '', 'low_rent': '', 'high_rent': '',
                    'utility': 'None', 'utility_display': 'Tenant-paid',
                    'report': report_date, 'market': market
                }

                import re as re3
                for t in tables:
                    for row in t:
                        if not row: continue
                        c = str(row[0] or '').strip()
                        v = str(row[1] or '').strip() if len(row) > 1 else ''
                        v2 = str(row[2] or '').strip() if len(row) > 2 else ''

                        # Name/address block
                        if '\n' in c and re3.search(r'(Avenue|Street|Road|Drive|Way|Blvd|Boulevard|Lane|Circle|Court|Pike|Place|Hwy|Highway|Ridge|Trail)', c):
                            lines = [l.strip() for l in c.split('\n') if l.strip()]
                            park['name'] = lines[0]
                            if len(lines) > 1: park['address'] = lines[1]
                            for ln in lines:
                                m = re3.search(r'([\w\s]+?),?\s+([A-Z]{2})\s+(\d{5})', ln)
                                if m:
                                    park['city'] = m.group(1).strip().rstrip(',')
                                    park['state'] = m.group(2)
                                    park['zip'] = m.group(3)
                                ph = re3.search(r'\d{3}[-.\s]\d{3}[-.\s]\d{4}', ln)
                                if ph and not park['phone']: park['phone'] = ph.group()

                        # Site info
                        if 'Total Sites' in c and '\n' in c:
                            keys = c.split('\n'); vals = v.split('\n')
                            for ki, kk in enumerate(keys):
                                vv = vals[ki].strip() if ki < len(vals) else ''
                                if 'Total Sites' in kk: park['spaces'] = vv
                                if 'Occupied Percent' in kk: park['occupancy_pct'] = vv

                        # Utility
                        if 'Water\nSewer\nTrash\nCable\nLawn' in c and '\n' in v:
                            svc_names = ['Water','Sewer','Trash','Cable','Lawn']
                            included = v.split('\n')
                            abbrevs = ['W','S','T','C','L']
                            parts = [abbrevs[k] for k,inc in enumerate(included) if inc.strip()=='Yes' and k<len(svc_names)]
                            park['utility'] = '/'.join(parts) if parts else 'None'
                            full_parts = [svc_names[k] for k,inc in enumerate(included) if inc.strip()=='Yes' and k<len(svc_names)]
                            park['utility_display'] = (', '.join(full_parts) + ' incl.') if full_parts else 'Tenant-paid'

                        # Rent
                        if c in ['All Homesites','All Sites','55+ Homesites','Senior Homesites','All Age Homesites']:
                            if v: park['avg_rent'] = v.replace('$','').replace(',','').strip()
                            if v2: park['adj_avg_rent'] = v2.replace('$','').replace(',','').strip()

                # Fallback rent from nearby pages
                if park['name'] and not park['avg_rent']:
                    for k in range(max(0,i-20), min(len(pdf.pages), i+5)):
                        cp_text = pdf.pages[k].extract_text() or ''
                        if park['name'][:12] in cp_text:
                            for ln in cp_text.split('\n'):
                                if park['name'][:12] in ln:
                                    rents = re3.findall(r'\$[\d,]+', ln)
                                    if len(rents) >= 3:
                                        park['low_rent'] = rents[0].replace('$','').replace(',','')
                                        park['high_rent'] = rents[1].replace('$','').replace(',','')
                                        park['avg_rent'] = rents[2].replace('$','').replace(',','')
                                    elif rents:
                                        park['avg_rent'] = rents[0].replace('$','').replace(',','')
                                    break
                            break

                if park['name'] and park['spaces']:
                    parks.append(park)

        if not parks:
            return jsonify({"success": False, "error": "No parks found in PDF. Make sure this is a JLT Market Report."}), 400

        # Upload to database
        db = load_jlt_db()
        added, updated = 0, 0
        for np in parks:
            found = False
            for idx, ep in enumerate(db):
                score = match_score(np['name'], ep['name'], np.get('city',''), ep.get('city',''))
                if score >= 85:
                    db[idx] = np
                    updated += 1
                    found = True
                    break
            if not found:
                db.append(np)
                added += 1
        save_jlt_db(db)

        return jsonify({"success": True, "parsed": len(parks), "added": added, "updated": updated, "total": len(db), "market": market})

    except Exception as e:
        return jsonify({"success": False, "error": str(e), "trace": traceback.format_exc()}), 500
